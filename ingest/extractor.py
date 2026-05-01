"""
Text extraction from PDF, DOCX, DOC, and XLSX files.

Extraction strategy (tiered):
    1. pdfplumber for embedded-text PDFs (fast, reliable).
    2. PaddleOCR 2.9.1 fallback for scanned pages or SAI Global watermarked PDFs
       where pdfplumber returns empty strings despite having visible content.
    3. python-docx for DOCX/DOCM files (paragraphs + tables).
    4. openpyxl for XLSX/XLSM/XLS files (extracts all cell values).
    5. subprocess antiword/catdoc for legacy .doc files.

PaddleOCR API: 2.9.1 uses ocr.ocr() with list-based output.
PaddlePaddle: 2.6.2 (GPU on Colab, CPU on Windows).

CHANGES (v2.2 — 30 Apr 2026):
  - PDF: Always extract tables (not just when text < threshold). Tables
    are appended as markdown after the page text so both are captured.
  - DOCX: Tables now formatted as markdown with header row + separator.
    Table position preserved relative to paragraphs (interleaved).
  - XLSX: First row treated as headers, remaining rows formatted as
    markdown table. Much better structure for LLM comprehension.
  - All table formatting uses _format_table_as_markdown() helper.
"""

import logging
import os
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


@dataclass
class ExtractionResult:
    """Result of extracting text from a single file."""
    file_path: str
    text: str
    page_count: int
    method: str  # "pdfplumber", "paddleocr", "docx", "xlsx", "doc", "skipped"
    ocr_pages: int = 0
    error: Optional[str] = None


class TextExtractor:
    """
    Extracts text from PDF, DOCX, DOC, and XLSX files using a tiered strategy.

    For PDFs: pdfplumber first, PaddleOCR fallback for pages with
    insufficient text (below min_text_threshold). This handles both
    scanned documents and SAI Global watermarked Australian Standards
    where encoded text streams confuse pdfplumber.

    For DOCX/DOCM: python-docx paragraph + table extraction.
    For XLSX/XLSM/XLS: openpyxl cell value extraction across all sheets.
    For DOC: antiword subprocess fallback.
    """

    SUPPORTED_EXTENSIONS = {
        ".pdf", ".docx", ".docm",
        ".xlsx", ".xlsm", ".xls",
        ".doc",
    }

    # Known DRM patterns in PDF metadata or content
    DRM_INDICATORS = [
        "sai global", "techstreet", "licensed to",
        "copyright", "not for resale", "single user licence",
    ]

    def __init__(
        self,
        enable_ocr: bool = True,
        ocr_timeout: int = 10,
        min_text_threshold: int = 10,
        ocr_confidence: float = 0.5,
    ):
        self._enable_ocr = enable_ocr
        self._ocr_timeout = ocr_timeout
        self._min_text_threshold = min_text_threshold
        self._ocr_confidence = ocr_confidence
        self._ocr_engine = None

    # ------------------------------------------------------------------
    # OCR engine — lazy load so non-OCR runs don't pay the import cost
    # ------------------------------------------------------------------

    def _get_ocr_engine(self):
        """Load PaddleOCR 2.9.1 on first use."""
        if self._ocr_engine is None:
            from paddleocr import PaddleOCR

            self._ocr_engine = PaddleOCR(
                use_angle_cls=True,
                lang="en",
                show_log=False,
                use_gpu=True,
            )
            logger.info("PaddleOCR 2.9.1 engine loaded (GPU enabled).")
        return self._ocr_engine

    # ------------------------------------------------------------------
    # Table formatting helper
    # ------------------------------------------------------------------

    @staticmethod
    def _format_table_as_markdown(rows: list[list]) -> str:
        """
        Convert a list of rows (each row = list of cell strings) into
        a markdown table with header row and separator.

        If the table has fewer than 2 rows, falls back to pipe-separated.

        Example output:
            | Header 1 | Header 2 | Header 3 |
            |----------|----------|----------|
            | Cell A   | Cell B   | Cell C   |
            | Cell D   | Cell E   | Cell F   |
        """
        if not rows:
            return ""

        # Clean cells: convert None to empty string, strip whitespace
        cleaned = []
        for row in rows:
            cleaned_row = [str(c).strip() if c is not None else "" for c in row]
            # Skip completely empty rows
            if any(cell for cell in cleaned_row):
                cleaned.append(cleaned_row)

        if not cleaned:
            return ""

        # Normalise column count (pad short rows)
        max_cols = max(len(r) for r in cleaned)
        for row in cleaned:
            while len(row) < max_cols:
                row.append("")

        # Build markdown table
        lines = []

        # Header row (first row)
        header = cleaned[0]
        lines.append("| " + " | ".join(header) + " |")

        # Separator row
        lines.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows
        for row in cleaned[1:]:
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)

    # ------------------------------------------------------------------
    # DRM detection
    # ------------------------------------------------------------------

    def _check_drm(self, file_path: str) -> bool:
        """
        Check if a PDF is DRM-protected by examining metadata and first page.
        Returns True if DRM detected.
        """
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                meta = pdf.metadata or {}
                meta_str = " ".join(str(v).lower() for v in meta.values() if v)
                for indicator in self.DRM_INDICATORS:
                    if indicator in meta_str:
                        return True

                if hasattr(pdf, 'is_encrypted') and pdf.is_encrypted:
                    return True

        except Exception:
            pass
        return False

    # ------------------------------------------------------------------
    # PDF extraction
    # ------------------------------------------------------------------

    def _extract_pdf(self, file_path: str) -> ExtractionResult:
        """
        Extract text from PDF using pdfplumber, with OCR fallback.

        CHANGED (v2.2): Tables are now ALWAYS extracted from every page,
        not just when extract_text() returns insufficient text. Both
        regular text and table text are combined per page.
        """
        import pdfplumber

        if self._check_drm(file_path):
            filename = Path(file_path).name
            logger.warning("  DRM detected: %s — extracting what we can", filename)

        pages_text = []
        ocr_page_count = 0
        total_pages = 0

        try:
            with pdfplumber.open(file_path) as pdf:
                total_pages = len(pdf.pages)

                for page_num, page in enumerate(pdf.pages):
                    page_parts = []

                    # --- 1. Extract regular text ---
                    text = ""
                    try:
                        text = (page.extract_text() or "").strip()
                    except Exception as e:
                        logger.debug("pdfplumber failed on page %d: %s", page_num, e)

                    if text:
                        page_parts.append(text)

                    # --- 2. ALWAYS extract tables (new in v2.2) ---
                    try:
                        tables = page.extract_tables()
                        if tables:
                            for table in tables:
                                if table and len(table) > 0:
                                    md_table = self._format_table_as_markdown(table)
                                    if md_table:
                                        page_parts.append(md_table)
                    except Exception as e:
                        logger.debug("Table extraction failed on page %d: %s", page_num, e)

                    # --- 3. If still no content, try OCR ---
                    page_text = "\n\n".join(page_parts).strip()

                    if len(page_text) < self._min_text_threshold and self._enable_ocr:
                        ocr_text = self._ocr_page(file_path, page_num)
                        if ocr_text and len(ocr_text) > len(page_text):
                            page_text = ocr_text
                            ocr_page_count += 1

                    if page_text:
                        pages_text.append(page_text)

        except Exception as e:
            error_str = str(e).lower()
            if "encrypted" in error_str or "password" in error_str:
                logger.warning("  DRM/encrypted PDF: %s", file_path)
                return ExtractionResult(
                    file_path=file_path, text="", page_count=0,
                    method="skipped", error=f"DRM/encrypted: {e}",
                )
            logger.error("PDF extraction failed for %s: %s", file_path, e)
            return ExtractionResult(
                file_path=file_path, text="", page_count=0,
                method="skipped", error=str(e),
            )

        combined = "\n\n".join(pages_text)
        method = "pdfplumber" if ocr_page_count == 0 else "pdfplumber+paddleocr"

        return ExtractionResult(
            file_path=file_path,
            text=combined,
            page_count=total_pages,
            method=method,
            ocr_pages=ocr_page_count,
        )

    def _ocr_page(self, pdf_path: str, page_num: int) -> str:
        """OCR a single PDF page using PaddleOCR 2.9.1 (ocr.ocr() API)."""
        try:
            from pdf2image import convert_from_path

            images = convert_from_path(
                pdf_path,
                first_page=page_num + 1,
                last_page=page_num + 1,
                dpi=200,
            )
            if not images:
                return ""

            import numpy as np

            img_array = np.array(images[0])
            ocr = self._get_ocr_engine()

            result = ocr.ocr(img_array, cls=True)

            if not result or not result[0]:
                return ""

            lines = []
            for line in result[0]:
                if len(line) >= 2:
                    text_conf = line[1]
                    if isinstance(text_conf, tuple) and len(text_conf) >= 2:
                        text, conf = text_conf[0], text_conf[1]
                        if conf >= self._ocr_confidence:
                            lines.append(text)

            return "\n".join(lines)

        except Exception as e:
            logger.warning("OCR failed for %s page %d: %s", pdf_path, page_num, e)
            return ""

    # ------------------------------------------------------------------
    # DOCX / DOCM extraction
    # ------------------------------------------------------------------

    def _extract_docx(self, file_path: str) -> ExtractionResult:
        """
        Extract text from DOCX/DOCM using python-docx.

        CHANGED (v2.2): Tables are now formatted as markdown with header
        row and separator. Table position is interleaved with paragraphs
        by walking the document body XML element order, so content appears
        in the same order as the original document.
        """
        try:
            from docx import Document
            from docx.table import Table as DocxTable
            from docx.text.paragraph import Paragraph

            doc = Document(file_path)
            parts = []

            # Walk the document body in element order to interleave
            # paragraphs and tables correctly.
            for element in doc.element.body:
                tag = element.tag.split("}")[-1]  # strip namespace

                if tag == "p":
                    # It's a paragraph
                    para = Paragraph(element, doc)
                    text = para.text.strip()
                    if text:
                        parts.append(text)

                elif tag == "tbl":
                    # It's a table — format as markdown
                    table = DocxTable(element, doc)
                    rows = []
                    for row in table.rows:
                        cells = [cell.text.strip() for cell in row.cells]
                        rows.append(cells)

                    md_table = self._format_table_as_markdown(rows)
                    if md_table:
                        parts.append(md_table)

            combined = "\n\n".join(parts)

            return ExtractionResult(
                file_path=file_path,
                text=combined,
                page_count=1,
                method="docx",
            )

        except Exception as e:
            logger.error("DOCX extraction failed for %s: %s", file_path, e)
            return ExtractionResult(
                file_path=file_path, text="", page_count=0,
                method="skipped", error=str(e),
            )

    # ------------------------------------------------------------------
    # DOC extraction (legacy Word format)
    # ------------------------------------------------------------------

    def _extract_doc(self, file_path: str) -> ExtractionResult:
        """Extract text from legacy .doc files using antiword or catdoc."""
        text = ""

        try:
            result = subprocess.run(
                ["antiword", file_path],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode == 0 and result.stdout.strip():
                text = result.stdout.strip()
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

        if not text:
            try:
                result = subprocess.run(
                    ["catdoc", file_path],
                    capture_output=True, text=True, timeout=30,
                )
                if result.returncode == 0 and result.stdout.strip():
                    text = result.stdout.strip()
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if text:
            return ExtractionResult(
                file_path=file_path, text=text,
                page_count=1, method="doc",
            )

        logger.warning("No .doc converter available for %s (install antiword or catdoc)", file_path)
        return ExtractionResult(
            file_path=file_path, text="", page_count=0,
            method="skipped", error="No .doc converter (antiword/catdoc not found)",
        )

    # ------------------------------------------------------------------
    # XLSX / XLSM / XLS extraction
    # ------------------------------------------------------------------

    def _extract_xlsx(self, file_path: str) -> ExtractionResult:
        """
        Extract text from Excel files using openpyxl.

        CHANGED (v2.2): First row is treated as headers. Remaining rows
        are formatted as a markdown table per sheet. This preserves column
        relationships that were lost in the old pipe-separated format.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            parts = []
            sheet_count = len(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = []

                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    # Keep row if it has at least one non-empty cell
                    if any(cell for cell in cells):
                        all_rows.append(cells)

                if all_rows:
                    parts.append(f"**Sheet: {sheet_name}**")
                    md_table = self._format_table_as_markdown(all_rows)
                    if md_table:
                        parts.append(md_table)

            wb.close()
            combined = "\n\n".join(parts)

            return ExtractionResult(
                file_path=file_path,
                text=combined,
                page_count=sheet_count if parts else 0,
                method="xlsx",
            )

        except Exception as e:
            logger.error("XLSX extraction failed for %s: %s", file_path, e)
            return ExtractionResult(
                file_path=file_path, text="", page_count=0,
                method="skipped", error=str(e),
            )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract(self, file_path: str) -> ExtractionResult:
        """
        Extract text from a single file.

        Args:
            file_path: Absolute path to the file.

        Returns:
            ExtractionResult with extracted text, page count, method used.
        """
        ext = Path(file_path).suffix.lower()

        if ext == ".pdf":
            return self._extract_pdf(file_path)
        elif ext in (".docx", ".docm"):
            return self._extract_docx(file_path)
        elif ext == ".doc":
            return self._extract_doc(file_path)
        elif ext in (".xlsx", ".xlsm", ".xls"):
            return self._extract_xlsx(file_path)
        else:
            return ExtractionResult(
                file_path=file_path, text="", page_count=0,
                method="skipped", error=f"Unsupported extension: {ext}",
            )

    def is_supported(self, file_path: str) -> bool:
        """Check if the file type is supported for extraction."""
        return Path(file_path).suffix.lower() in self.SUPPORTED_EXTENSIONS
